import openpyxl as vb
from win32com.client import Dispatch
row = 11
path2 = r"F:/excel/尺码表模板.xlsx"
path3 = r"F:/excel/尺码表最终.xlsx"
def get_alldata(row):
    path1 = r"F:/excel/尺码表.xlsx"
    wb1 = vb.load_workbook(path1,data_only=True)
    sheet1 = wb1['总表']
    a = sheet1['A'+str(row):'K'+str(row)]
    # print(a)
    list1=[]
    for date in a:
        for box in date:
            if box.value == None:
                continue
            else:
                list1.append(box.value)
    return list1

def get_list (row):
    alllist=[]
    for r in range(2,row+1):
        alllist.append(get_alldata(r))
    return alllist

def get_dic(alllist):
    dic ={}
    for onelist in alllist:
        dic[onelist[0]] = onelist[1:]
    return dic

def copy_excel (row):
    list1 = get_list(row)
    wb = vb.load_workbook(path2)
    wb.save(path3)
    cpwb = vb.load_workbook(path3)
    for num in range(0,row-1):
        column = str(list1[num][3])+'列'
        name = str(list1[num][0])
        copy_st = cpwb.copy_worksheet(cpwb[column])
        copy_st.title = name

    for rm in range(3,8):
        cpwb.remove(cpwb[str(rm)+'列'])
    cpwb.save(path3)

def copydata(dic):
    size_table = [['F', 'C', '3'], ['XS', 'C',  '4'], ['S','C',  '5'],['M' ,'C', '6'],['L' , 'C', '7'],['XL','C', '8'],['XXL' ,'C', '9'],['XXXL','C', '10']]
    zzwb = vb.load_workbook(path3)
    for k, v in dic.items():
        sheet = zzwb[k]
        size = v[1]
        size_data = v[3:]
        for size1 in size_table:
            if size1[0] == size:
                for value2 in sheet['C'+size1[2]:'I'+size1[2]]:
                    for value1,size2 in zip(value2,size_data):
                        value1.value = size2
    zzwb.save(path3)

def just_open():
 xlApp = Dispatch("Excel.Application")
 xlApp.Visible = False
 xlBook = xlApp.Workbooks.Open(path3)
 xlBook.Save()
 xlBook.Close()

def copy_value():
    zzwb = vb.load_workbook(path3,data_only=True)
    for sheet in zzwb:
        for value3 in sheet['C3':'I10']:
            for value4 in value3:
                if value4.value == '#VALUE!':
                    value4.value = '/'
    zzwb.save(path3)

def delete_coloum(dic):
    size_table=[['F' ,'C', '3'],['XS' ,'C',  '4'],['S' ,'C',  '5'],['M' ,'C', '6'],['L' , 'C', '7'],['XL','C', '8'],['XXL' ,'C', '9'],['XXXL','C', '10']]
    zzwb = vb.load_workbook(path3)
    for k, v in dic.items():
        size_range = v[0]  # 获取尺码范围
        sheet = zzwb[k]
        size = v[1]
        for size1 in size_table:
            if size1[0] == size:
                idx1 = int(size1[2]) + int(size_range)
                if idx1 == 4:
                    sheet.delete_rows(idx=idx1, amount=11-idx1)
                elif idx1 >4:
                    sheet.delete_rows(idx=idx1, amount=11-idx1)
                    sheet.delete_rows(idx=3, amount=int(size1[2])-3)  #删除多余的行列

    zzwb.save(path3)

def merger(dic):
    zzwb = vb.load_workbook(path3)
    for k,v in dic.items():
        sheet = zzwb[k]
        lenxl = int(v[2])
        highxl =int(v[0])
        list1 =['C','D','E','F','G','H','I','J']
        lenstr=list1[lenxl]
        list2 =[3,4,5,6,7,8,9,10,11]
        highstr=list2[highxl]
        sheet.merge_cells('A'+str(highstr)+':'+lenstr+str(highstr+1))
    zzwb.save(path3)





get_alldata(row)
copy_excel(row)
t = get_list(row)
dic = get_dic(t)
copydata(dic)
just_open()
copy_value()
delete_coloum(dic)
merger(dic)




#
# if __name__ == '__main__':
#




"""
作者：zelin
日期：2021年11年04年
"""
